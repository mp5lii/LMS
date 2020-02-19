#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import time

import pyodbc
import wx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter


class TabUsers(wx.Panel):
    
    def __init__(self, parent, size=wx.Size(750, 500), style=wx.TAB_TRAVERSAL, name=wx.EmptyString):
        wx.Panel.__init__(self, parent, id=wx.ID_ANY, pos=wx.DefaultPosition, size=size, style=style, name=name)
        
        font = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString)
        # ######################################界面元素###################################### #
        # 班级静态文本框
        st_class = wx.StaticText(self, wx.ID_ANY, u"班级：", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_LEFT)
        st_class.Wrap(-1)
        st_class.SetFont(font)
        # 在数据库中查找用户班级
        conn = pyodbc.connect("DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb" % os.getcwd())
        cursor = conn.cursor()
        cursor.execute("SELECT user_class FROM users GROUP BY user_class")
        rows = cursor.fetchall()
        classes = ["请选择"]
        for row in rows:
            if row[0][-2:] != "毕业":
                classes.append(row[0])
        cursor.close()
        conn.close()
        self._choice_class = wx.Choice(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, classes)
        self._choice_class.SetSelection(0)
        self._choice_class.SetFont(font)
        # 用户列表
        style = wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.LC_HRULES | wx.LC_VRULES | wx.LC_EDIT_LABELS | wx.VSCROLL
        self._lc_users = wx.ListCtrl(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, style)
        self._lc_users.SetFont(font)
        self._lc_users.InsertColumn(0, '', wx.LIST_FORMAT_CENTER, width=2)
        self._lc_users.InsertColumn(1, '序号', wx.LIST_FORMAT_CENTER, width=50)
        self._lc_users.InsertColumn(2, 'ID卡号', wx.LIST_FORMAT_CENTER, width=120)
        self._lc_users.InsertColumn(3, '姓名', wx.LIST_FORMAT_CENTER, width=130)
        self._lc_users.InsertColumn(4, '照片', wx.LIST_FORMAT_CENTER, width=150)
        self._lc_users.InsertColumn(5, '在借图书', wx.LIST_FORMAT_CENTER, width=100)
        # 导入用户按钮
        self._btn_import_users = wx.Button(self, wx.ID_ANY, u"导入用户", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_import_users.SetFont(font)
        # 覆盖用户复选框
        self._chk_cover_users_data = wx.CheckBox(self, wx.ID_ANY, u"覆盖用户", wx.DefaultPosition, wx.DefaultSize, 0)
        self._chk_cover_users_data.SetFont(font)
        # 导出用户按钮
        self._btn_export_users = wx.Button(self, wx.ID_ANY, u"导出用户", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_export_users.SetFont(font)
        # 更新用户班级按钮
        self._btn_update_classes = wx.Button(self, wx.ID_ANY, u"更新用户班级", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_update_classes.SetFont(font)
        # 用户借书列表c
        self._lc_user_books = wx.ListCtrl(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                          wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.LC_HRULES | wx.LC_VRULES)
        self._lc_user_books.SetFont(font)
        self._lc_user_books.InsertColumn(0, '', wx.LIST_FORMAT_CENTER, width=2)
        self._lc_user_books.InsertColumn(1, '序号', wx.LIST_FORMAT_CENTER, width=45)
        self._lc_user_books.InsertColumn(2, 'ISBN', wx.LIST_FORMAT_CENTER, width=100)
        self._lc_user_books.InsertColumn(3, '名称', wx.LIST_FORMAT_CENTER, width=350)
        self._lc_user_books.InsertColumn(4, '摆放位置', wx.LIST_FORMAT_CENTER, width=130)
        # 删除借书记录按钮
        self._btn_delete_user_book = wx.Button(self, wx.ID_ANY, u"删除记录", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_delete_user_book.SetFont(font)
        # 添加借书记录按钮
        self._btn_add_user_book = wx.Button(self, wx.ID_ANY, u"添加记录", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_add_user_book.SetFont(font)
        # ######################################界面元素###################################### #
        
        # ######################################布局信息###################################### #
        bs_user = wx.BoxSizer(wx.VERTICAL)
        bs_user.Add(self._btn_import_users, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_user.Add(self._chk_cover_users_data, 0, wx.BOTTOM | wx.ALIGN_CENTER_HORIZONTAL, 20)
        bs_user.Add(self._btn_export_users, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_user.Add(self._btn_update_classes, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_book = wx.BoxSizer(wx.VERTICAL)
        bs_book.Add(self._btn_add_user_book, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_book.Add(self._btn_delete_user_book, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        
        gbs = wx.GridBagSizer(0, 0)
        gbs.SetFlexibleDirection(wx.BOTH)
        gbs.SetNonFlexibleGrowMode(wx.FLEX_GROWMODE_SPECIFIED)
        gbs.Add(st_class, wx.GBPosition(0, 0), wx.GBSpan(1, 1),
                wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL | wx.TOP, 20)
        gbs.Add(self._choice_class, wx.GBPosition(1, 0), wx.GBSpan(1, 1), wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        gbs.Add(self._lc_users, wx.GBPosition(0, 1), wx.GBSpan(2, 1),
                wx.ALL | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL, 5)
        gbs.Add(bs_user, wx.GBPosition(0, 2), wx.GBSpan(2, 1),
                wx.EXPAND | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL, 5)
        gbs.Add((0, 0), wx.GBPosition(2, 0), wx.GBSpan(1, 3),
                wx.EXPAND | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        gbs.Add(self._lc_user_books, wx.GBPosition(3, 0), wx.GBSpan(1, 2),
                wx.ALL | wx.EXPAND | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        gbs.Add(bs_book, wx.GBPosition(3, 2), wx.GBSpan(1, 1),
                wx.ALL | wx.ALIGN_BOTTOM | wx.ALIGN_CENTER_HORIZONTAL, 5)
        # ######################################布局信息###################################### #
        
        self.SetSizerAndFit(gbs)
        self.Layout()
        
        # Connect Events
        self._choice_class.Bind(wx.EVT_CHOICE, self._on_choice_class)
        self._lc_users.Bind(wx.EVT_LIST_ITEM_SELECTED, self._on_lc_users_selected)
        self._btn_import_users.Bind(wx.EVT_BUTTON, self._on_btn_import_users)
        self._btn_export_users.Bind(wx.EVT_BUTTON, self._on_btn_export_users)
        self._btn_update_classes.Bind(wx.EVT_BUTTON, self._on_btn_update_classes)
        self._btn_add_user_book.Bind(wx.EVT_BUTTON, self._on_btn_add_user_book)
        self._btn_delete_user_book.Bind(wx.EVT_BUTTON, self._on_btn_delete_user_book)
    
    def __del__(self):
        pass
    
    # Virtual event handlers, overide them in your derived class
    def _on_lc_users_selected(self, event):
        self._lc_user_books.DeleteAllItems()
        user_id = self._lc_users.GetItemText(self._lc_users.GetFirstSelected(), 2)
        books = self._lc_users.GetItemText(self._lc_users.GetFirstSelected(), 5)
        if books != "":
            # 新建数据库连接
            conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
            cursor = conn.cursor()
            cursor_tmp = conn.cursor()
            # 在数据库中查找用户姓名、照片和班级
            cursor.execute("SELECT book_isbn FROM records WHERE user_id = ?", user_id)
            rows = cursor.fetchall()
            if len(rows) > 0:
                for row in rows:
                    cursor_tmp.execute("SELECT book_name, book_position FROM books WHERE book_isbn = ?", row[0])
                    row_tmp = cursor_tmp.fetchone()
                    item_count = self._lc_user_books.GetItemCount()
                    index = self._lc_user_books.InsertItem(item_count, str(item_count))
                    # 序号
                    self._lc_user_books.SetItem(index, 1, str(item_count + 1))
                    # ISBN
                    self._lc_user_books.SetItem(index, 2, row[0])
                    # 图书名称
                    self._lc_user_books.SetItem(index, 3, row_tmp[0])
                    # 摆放位置
                    if row_tmp[1] is None:
                        self._lc_user_books.SetItem(index, 4, "")
                    else:
                        self._lc_user_books.SetItem(index, 4, row_tmp[1])
            cursor_tmp.close()
            cursor.close()
            conn.close()
        event.Skip()
    
    def _on_choice_class(self, event):
        idx = self._choice_class.GetSelection()
        class_name = self._choice_class.GetString(idx)
        self._lc_users.DeleteAllItems()
        # 数据库中查询该班级人员及借书情况
        conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
        cursor = conn.cursor()
        cursor_tmp = conn.cursor()
        # 在数据库中查找用户姓名、照片和班级
        cursor.execute("SELECT user_id, user_name, user_photo FROM users WHERE user_class = ?", class_name)
        while 1:
            row = cursor.fetchone()
            if not row:  # 数据库中有该用户
                break
            item_count = self._lc_users.GetItemCount()
            index = self._lc_users.InsertItem(item_count, str(item_count))
            # 用于查找
            self._lc_users.SetItem(index, 0, row[0])
            # 序号
            self._lc_users.SetItem(index, 1, str(item_count + 1))
            # ID
            self._lc_users.SetItem(index, 2, row[0])
            # 姓名
            self._lc_users.SetItem(index, 3, row[1])
            # 照片（可能为空）
            if row[2] is None:
                self._lc_users.SetItem(index, 4, '')
            else:
                self._lc_users.SetItem(index, 4, row[2])
            # 借书数量
            cursor_tmp.execute("SELECT book_isbn FROM records WHERE user_id = ?", row[0])
            row_tmp = cursor_tmp.fetchall()
            books_count = len(row_tmp)
            if books_count == 0:
                self._lc_users.SetItem(index, 5, '')
            else:
                self._lc_users.SetItem(index, 5, str(books_count) + u'本')
        # 关闭数据库
        cursor.close()
        cursor_tmp.close()
        conn.close()
        event.Skip()
    
    def _on_btn_import_users(self, event):
        file_dialog = wx.FileDialog(self, "选择要导入的用户数据文件",
                                    wildcard="Excel 文件(*.xlsx)|*.xlsx|Excel97-2003 文件(*.xls)|*.xls",
                                    style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)
        if file_dialog.ShowModal() == wx.ID_OK:
            # 导入操作进行时设置鼠标形状为箭头等待状态
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROWWAIT))
            # 记录是否覆盖数据状态
            b_cover_data = self._chk_cover_users_data.GetValue()
            # 获取待导入的文件
            file = file_dialog.GetPath()
            # 打开Excel文件
            conn_str = (r'DRIVER={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};DBQ=%s;' % file)
            conn_excel = pyodbc.connect(conn_str, autocommit=True)
            cursor_excel = conn_excel.cursor()
            cursor_excel_tmp = conn_excel.cursor()
            # 连接数据库
            conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
            cursor = conn.cursor()
            # 逐个表读取数据
            for worksheet in cursor_excel.tables():
                # 获取表名
                table_name = worksheet[2]
                # 读取表内容
                sel_str = (r'SELECT * FROM [%s]' % table_name)
                excel_rows = cursor_excel_tmp.execute(sel_str).fetchall()
                # 逐行读取
                for index, excel_row in enumerate(excel_rows):
                    # 行数据不完善（用户Id、用户名字或用户班级为空）时，跳过该行
                    if (excel_row[0] is None) or (excel_row[1] is None) or (excel_row[3] is None):
                        continue
                    # 在数据库中查找是否存在该Id的用户
                    cursor.execute("SELECT user_id FROM users WHERE user_id = ?", excel_row[0])
                    user = cursor.fetchone()
                    if user is None:  # 无此用户，则将该用户添加到数据库中
                        cursor.execute("""
                            INSERT INTO users(user_id, user_name, user_photo, user_class)
                            VALUES(?, ?, ?, ?)""", excel_row[0], excel_row[1], excel_row[2], excel_row[3])
                    elif b_cover_data:  # 有此用户，且需要覆盖原数据时，则更新该用户信息。（不能覆盖原数据时不进行操作）
                        cursor.execute("""
                            UPDATE users SET user_name=?, user_photo=?, user_class=?
                            WHERE user_id=?""", excel_row[1], excel_row[2], excel_row[3], excel_row[0])
                    conn.commit()
            # 关闭Excel文件
            cursor_excel_tmp.close()
            cursor_excel.close()
            conn_excel.close()
            # 关闭数据库
            cursor.close()
            conn.close()
            # 恢复鼠标形状
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROW))
            # 弹出提示框
            wx.MessageBox("导入完成！")
            # 更新显示当前班级用户
            wx.QueueEvent(self, wx.PyCommandEvent(wx.EVT_CHOICE))
        event.Skip()
    
    def _on_btn_export_users(self, event):
        default_dir = os.path.join(os.path.expanduser("~"), 'Desktop')
        default_file = u"导出用户" + time.strftime('%Y-%m-%d', time.localtime(time.time())) + r".xlsx"
        file_dialog = wx.FileDialog(self, u"导出用户数据", defaultDir=default_dir, defaultFile=default_file,
                                    wildcard=u"Excel 文件(*.xlsx)|",
                                    style=wx.FD_SAVE)
        if file_dialog.ShowModal() == wx.ID_OK:
            # 导处操作进行时设置鼠标形状为箭头等待状态
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROWWAIT))
            # 创建Excel文件
            work_book = Workbook()
            work_sheet = work_book.active
            work_sheet.title = u"导出用户"
            # 写表头
            row = ("", "ID", "姓名", "照片", "班级")
            work_sheet.append(row)
            for i in range(1, 6):
                cell = get_column_letter(i) + '1'
                work_sheet[cell].style = 'Headline 1'
                work_sheet[cell].fill = PatternFill("solid", fgColor="DDDDDD")
                work_sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            # 初始化
            class_name = u""
            start_row = 1
            end_row = 1
            # 连接数据库
            conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
            cursor = conn.cursor()
            # 检索数据库中users表
            cursor.execute("SELECT user_id, user_name, user_photo, user_class FROM users ORDER BY user_class DESC")
            rows = cursor.fetchall()
            for row in rows:
                # 获取一行数据
                user = [row[3], row[0], row[1], row[2], row[3]]
                if row[2] is None:  # 用户照片为空时，该位置设置为空字符串
                    user[3] = ""
                # 输出一行
                work_sheet.append(user)
                # 如果是新班级
                if class_name != row[3]:
                    # 合并上一班级单元格
                    work_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    class_cell = work_sheet["A" + str(start_row)]
                    class_cell.value = class_name
                    class_cell.alignment = Alignment(horizontal="center", vertical="center")
                    # 记录新班级开始行
                    start_row = end_row + 1
                    # 记录新班级名称
                    class_name = row[3]
                # 下一行
                end_row += 1
            # 合并最后一个班级
            work_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            class_cell = work_sheet["A" + str(start_row)]
            class_cell.value = class_name
            class_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 关闭数据库
            cursor.close()
            conn.close()
            # 保存Excel文件
            work_book.save(filename=file_dialog.GetPath())
            # 恢复鼠标形状
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROW))
            # 弹出提示框
            wx.MessageBox("导出完成！")
        event.Skip()
    
    @staticmethod
    def _on_btn_update_classes(self, event):
        # 连接数据库
        conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
        cursor = conn.cursor()
        # 将所有大班改为“*年8月大*毕业”
        year = time.strftime('%Y', time.localtime(time.time())) + u"年7月"
        cursor.execute("SELECT user_id, user_class FROM users WHERE user_class LIKE ?", u"大%")
        rows = cursor.fetchall()
        for row in rows:
            new_class = year + row[1][:] + "毕业"
            cursor.execute("UPDATE users SET user_class=? WHERE user_id=?", new_class, row[0])
            conn.commit()
        # 更改中班为大班
        cursor.execute("SELECT user_id, user_class FROM users WHERE user_class LIKE ?", u"中%")
        rows = cursor.fetchall()
        for row in rows:
            new_class = "大" + row[1][1:]
            cursor.execute("UPDATE users SET user_class=? WHERE user_id=?", new_class, row[0])
            conn.commit()
        # 更改小班为中班
        cursor.execute("SELECT user_id, user_class FROM users WHERE user_class LIKE ?", u"小%")
        rows = cursor.fetchall()
        for row in rows:
            new_class = "中" + row[1][1:]
            cursor.execute("UPDATE users SET user_class=? WHERE user_id=?", new_class, row[0])
            conn.commit()
        cursor.close()
        conn.close()
        # 弹出提示框
        wx.MessageBox("更新完成！")
        event.Skip()
    
    @staticmethod
    def _update_classes(self):
        # 连接数据库
        conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
        cursor = conn.cursor()
        # 将所有大班改为“*年8月大*毕业”
        year = time.strftime('%Y', time.localtime(time.time())) + u"年7月"
        cursor.execute("SELECT user_id, user_class FROM users WHERE user_class LIKE ?", u"大%")
        rows = cursor.fetchall()
        for row in rows:
            new_class = year + row[1][:] + "毕业"
            cursor.execute("UPDATE users SET user_class=? WHERE user_id=?", new_class, row[0])
            conn.commit()
        # 更改中班为大班
        cursor.execute("SELECT user_id, user_class FROM users WHERE user_class LIKE ?", u"中%")
        rows = cursor.fetchall()
        for row in rows:
            new_class = "大" + row[1][1:]
            cursor.execute("UPDATE users SET user_class=? WHERE user_id=?", new_class, row[0])
            conn.commit()
        # 更改小班为中班
        cursor.execute("SELECT user_id, user_class FROM users WHERE user_class LIKE ?", u"小%")
        rows = cursor.fetchall()
        for row in rows:
            new_class = "中" + row[1][1:]
            cursor.execute("UPDATE users SET user_class=? WHERE user_id=?", new_class, row[0])
            conn.commit()
        cursor.close()
        conn.close()
    
    def _on_btn_add_user_book(self, event):
        pass
    
    def _on_btn_delete_user_book(self, event):
        if self._lc_user_books.GetFirstSelected() != -1:
            user_id = self._lc_users.GetItemText(self._lc_users.GetFirstSelected(), 2)
            book_isbn = self._lc_user_books.GetItemText(self._lc_user_books.GetFirstSelected(), 2)
            # 新建数据库连接
            conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
            cursor = conn.cursor()
            # 数据库中查询该人员借书情况
            cursor.execute("DELETE FROM records WHERE user_id = ? AND book_isbn = ?", user_id, book_isbn)
            conn.commit()
            cursor.close()
            conn.close()
            self._lc_user_books.DeleteItem(self._lc_user_books.GetFirstSelected())
        event.Skip()


class TabBooks(wx.Panel):
    
    def __init__(self, parent, size=wx.Size(750, 500), style=wx.TAB_TRAVERSAL, name=wx.EmptyString):
        wx.Panel.__init__(self, parent, id=wx.ID_ANY, pos=wx.DefaultPosition, size=size, style=style, name=name)
        
        font = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString)
        
        # ######################################界面元素###################################### #
        self._lc_books = wx.ListCtrl(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                     wx.LC_HRULES | wx.LC_REPORT | wx.LC_VRULES | wx.VSCROLL)
        self._lc_books.SetFont(font)
        self._lc_books.InsertColumn(0, '', wx.LIST_FORMAT_CENTER, width=2)
        self._lc_books.InsertColumn(1, '序号', wx.LIST_FORMAT_CENTER, width=50)
        self._lc_books.InsertColumn(2, 'ISBN', wx.LIST_FORMAT_CENTER, width=120)
        self._lc_books.InsertColumn(3, '图书名称', wx.LIST_FORMAT_CENTER, width=250)
        self._lc_books.InsertColumn(4, '照片', wx.LIST_FORMAT_CENTER, width=120)
        self._lc_books.InsertColumn(5, '摆放位置', wx.LIST_FORMAT_CENTER, width=120)
        self._btn_import_books = wx.Button(self, wx.ID_ANY, u"导入图书", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_import_books.SetFont(font)
        self._chk_cover_books_data = wx.CheckBox(self, wx.ID_ANY, u"覆盖数据", wx.DefaultPosition, wx.DefaultSize, 0)
        self._chk_cover_books_data.SetFont(font)
        self._btn_export_books = wx.Button(self, wx.ID_ANY, u"导出图书", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_export_books.SetFont(font)
        self._btn_delete_books = wx.Button(self, wx.ID_ANY, u"删除图书", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_delete_books.SetFont(font)
        self._btn_clear_books = wx.Button(self, wx.ID_ANY, u"清空图书", wx.DefaultPosition, wx.DefaultSize, 0)
        self._btn_clear_books.SetFont(font)
        # ######################################界面元素###################################### #
        
        # ######################################布局信息###################################### #
        bs_operation = wx.BoxSizer(wx.VERTICAL)
        bs_operation.Add(self._btn_import_books, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_operation.Add(self._chk_cover_books_data, 0, wx.BOTTOM | wx.ALIGN_CENTER_HORIZONTAL, 20)
        bs_operation.Add(self._btn_export_books, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_operation.Add(self._btn_delete_books, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        bs_operation.Add(self._btn_clear_books, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        
        gbs = wx.GridBagSizer(0, 0)
        gbs.SetFlexibleDirection(wx.BOTH)
        gbs.SetNonFlexibleGrowMode(wx.FLEX_GROWMODE_SPECIFIED)
        gbs.Add(self._lc_books, wx.GBPosition(0, 0), wx.GBSpan(1, 1),
                wx.ALL | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL | wx.EXPAND, 5)
        gbs.Add(bs_operation, wx.GBPosition(0, 1), wx.GBSpan(1, 1), wx.EXPAND, 5)
        gbs.AddGrowableRow(0)
        
        self.SetSizerAndFit(gbs)
        # ######################################布局信息###################################### #
        
        self._get_books()
        
        # Connect Events
        self._btn_import_books.Bind(wx.EVT_BUTTON, self._on_btn_import_books)
        self._btn_export_books.Bind(wx.EVT_BUTTON, self._on_btn_export_books)
        self._btn_delete_books.Bind(wx.EVT_BUTTON, self._on_btn_delete_books)
        self._btn_clear_books.Bind(wx.EVT_BUTTON, self._on_btn_clear_books)
    
    def __del__(self):
        pass
    
    def _on_btn_import_books(self, event):
        file_dialog = wx.FileDialog(self, u"选择要导入的图书数据文件",
                                    wildcard=u"Excel 文件(*.xlsx)|*.xlsx|Excel97-2003 文件(*.xls)|*.xls",
                                    style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)
        if file_dialog.ShowModal() == wx.ID_OK:
            # 导入操作进行时设置鼠标形状为箭头等待状态
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROWWAIT))
            # 记录是否覆盖数据状态
            b_cover_data = self._chk_cover_books_data.GetValue()
            # 获取待导入的文件
            file = file_dialog.GetPath()
            # 打开Excel文件
            conn_str = (r'DRIVER={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};DBQ=%s;' % file)
            conn_excel = pyodbc.connect(conn_str, autocommit=True)
            cursor_excel = conn_excel.cursor()
            cursor_excel_tmp = conn_excel.cursor()
            # 连接数据库
            conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
            cursor = conn.cursor()
            # 逐个表读取数据
            for worksheet in cursor_excel.tables():
                # 获取表名
                table_name = worksheet[2]
                # 读取表内容
                sel_str = (r'SELECT * FROM [%s]' % table_name)
                excel_rows = cursor_excel_tmp.execute(sel_str).fetchall()
                # 逐行读取
                for index, excel_row in enumerate(excel_rows):
                    # 行数据不完善（图书ISBN或图书名字为空）时，跳过该行
                    if (excel_row[0] is None) or (excel_row[1] is None):
                        continue
                    # 在数据库中查找是否存在该图书
                    cursor.execute("SELECT book_isbn FROM books WHERE book_isbn = ?", excel_row[0])
                    book = cursor.fetchone()
                    if book is None:  # 无此图书，则将该图书添加到数据库中
                        cursor.execute("""
                            INSERT INTO books(book_isbn, book_name, book_photo, book_position)
                            VALUES(?, ?, ?, ?)""", excel_row[0], excel_row[1], excel_row[2], excel_row[3])
                    elif b_cover_data:  # 有此图书，且需要覆盖原数据时，则更新该图书信息。（不能覆盖原数据时不进行操作）
                        cursor.execute("""
                            UPDATE books SET book_name=?, book_photo=?, book_position=?
                            WHERE book_isbn=?""", excel_row[1], excel_row[2], excel_row[3], excel_row[0])
                    conn.commit()
            # 关闭Excel文件
            cursor_excel_tmp.close()
            cursor_excel.close()
            conn_excel.close()
            # 关闭数据库
            cursor.close()
            conn.close()
            # 恢复鼠标形状
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROW))
            # 弹出提示框
            wx.MessageBox(u"导入完成！")
            # 更新显示图书
            self._get_books()
        event.Skip()
    
    def _on_btn_export_books(self, event):
        default_dir = os.path.join(os.path.expanduser("~"), 'Desktop')
        default_file = u"导出图书" + time.strftime('%Y-%m-%d', time.localtime(time.time())) + r".xlsx"
        file_dialog = wx.FileDialog(self, u"导出图书数据", defaultDir=default_dir, defaultFile=default_file,
                                    wildcard="Excel 文件(*.xlsx)|", style=wx.FD_SAVE)
        if file_dialog.ShowModal() == wx.ID_OK:
            # 导处操作进行时设置鼠标形状为箭头等待状态
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROWWAIT))
            # 创建Excel文件
            work_book = Workbook()
            work_sheet = work_book.active
            work_sheet.title = u"导出图书"
            # 写表头
            row = ("ISBN", "图书名称", "图书照片", "摆放位置")
            work_sheet.append(row)
            for i in range(1, 5):
                cell = get_column_letter(i) + '1'
                work_sheet[cell].style = 'Headline 1'
                work_sheet[cell].fill = PatternFill("solid", fgColor="DDDDDD")
                work_sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            # 连接数据库
            conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
            cursor = conn.cursor()
            # 检索数据库中users表
            cursor.execute("SELECT book_isbn, book_name, book_photo, book_position FROM books")
            rows = cursor.fetchall()
            for row in rows:
                # 获取一行数据
                user = [row[0], row[1], row[2], row[3]]
                if row[2] is None:  # 用户照片为空时，该位置设置为空字符串
                    user[2] = ""
                if row[3] is None:
                    user[3] = ""
                # 输出一行
                work_sheet.append(user)
            # 关闭数据库
            cursor.close()
            conn.close()
            # 保存Excel文件
            work_book.save(filename=file_dialog.GetPath())
            # 恢复鼠标形状
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROW))
            # 弹出提示框
            wx.MessageBox("导出完成！")
        event.Skip()
    
    def _on_btn_delete_books(self, event):
        # 连接数据库
        conn = pyodbc.connect('DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb' % os.getcwd())
        cursor = conn.cursor()
        # 获取选中图书
        delete_item = []
        idx = self._lc_books.GetFirstSelected()
        while idx != wx.NOT_FOUND:
            delete_item.append(idx)
            idx = self._lc_books.GetNextSelected(idx)
        # 逆序删除选中图书
        delete_item.reverse()
        for idx in delete_item:
            book_isbn = self._lc_books.GetItemText(idx, 2)
            cursor.execute("DELETE FROM books WHERE book_isbn = ?", book_isbn)
            cursor.commit()
            self._lc_books.DeleteItem(idx)
        # 关闭数据库
        cursor.close()
        conn.close()
        event.Skip()
    
    def _on_btn_clear_books(self, event):
        # 导入操作进行时设置鼠标形状为箭头等待状态
        self.SetCursor(wx.Cursor(wx.CURSOR_ARROWWAIT))
        # 连接数据库
        conn = pyodbc.connect("DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb" % os.getcwd())
        cursor = conn.cursor()
        # 删除数据库中所有图书
        cursor.execute("DELETE * FROM books")
        cursor.commit()
        # 关闭数据库
        cursor.close()
        conn.close()
        # 弹出提示框
        wx.MessageBox(u"清空图书完成！")
        # 删除图书列表中全部内容
        self._lc_books.DeleteAllItems()
        event.Skip()
    
    def _get_books(self):
        self._lc_books.DeleteAllItems()
        # 在数据库中查找图书数据
        conn = pyodbc.connect("DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s\\LMSdb.accdb" % os.getcwd())
        cursor = conn.cursor()
        cursor.execute("SELECT book_isbn, book_name, book_photo, book_position FROM books")
        rows = cursor.fetchall()
        idx = 0
        for row in rows:
            if row[2] is None:
                book_photo = ""
            else:
                book_photo = row[2]
            if row[3] is None:
                book_position = ""
            else:
                book_position = row[3]
            book = ["", str(idx + 1), row[0], row[1], book_photo, book_position]
            self._lc_books.Append(book)
            idx += 1
        cursor.close()
        conn.close()


# class TabStatistics(wx.Panel):
#
#     def __init__(self, parent, id=wx.ID_ANY, pos=wx.DefaultPosition, size=wx.Size(750, 500), style=wx.TAB_TRAVERSAL,
#                  name=wx.EmptyString):
#         wx.Panel.__init__(self, parent, id=id, pos=pos, size=size, style=style, name=name)
#
#         b_sizerF = wx.BoxSizer(wx.VERTICAL)
#         font = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString)
#
#         gb_sizer = wx.GridBagSizer(0, 0)
#         gb_sizer.SetFlexibleDirection(wx.BOTH)
#         gb_sizer.SetNonFlexibleGrowMode(wx.FLEX_GROWMODE_SPECIFIED)
#
#         m_choiceChoices = [u"最常借的"]
#         self.m_choice = wx.Choice(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, m_choiceChoices, 0)
#         self.m_choice.SetSelection(0)
#         self.m_choice.SetFont(font)
#
#         gb_sizer.Add(self.m_choice, wx.GBPosition(0, 0), wx.GBSpan(1, 1),
#                      wx.ALL | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL, 5)
#
#         m_choice3Choices = [u"1", u"2", u"3", u"4", u"5", u"6", u"7", u"8", u"9", u"10"]
#         self.m_choice3 = wx.Choice(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, m_choice3Choices, 0)
#         self.m_choice3.SetSelection(4)
#         self.m_choice3.SetFont(font)
#
#         gb_sizer.Add(self.m_choice3, wx.GBPosition(0, 1), wx.GBSpan(1, 1),
#                      wx.ALL | wx.ALIGN_CENTER_HORIZONTAL | wx.ALIGN_CENTER_VERTICAL, 5)
#
#         self._list_book_statistics = wx.ListCtrl(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
#                                                  wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.LC_HRULES | wx.LC_VRULES)
#         self._list_book_statistics.SetFont(font)
#
#         gb_sizer.Add(self._list_book_statistics, wx.GBPosition(0, 3), wx.GBSpan(1, 1),
#                      wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER_HORIZONTAL, 5)
#
#         self.m_staticText = wx.StaticText(self, wx.ID_ANY, u"本书", wx.DefaultPosition, wx.DefaultSize, 0)
#         self.m_staticText.Wrap(-1)
#
#         self.m_staticText.SetFont(font)
#
#         gb_sizer.Add(self.m_staticText, wx.GBPosition(0, 2), wx.GBSpan(1, 1),
#                      wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER_HORIZONTAL, 5)
#
#         b_sizerF.Add(gb_sizer, 1, wx.ALIGN_CENTER_HORIZONTAL, 5)
#
#         self.SetSizer(b_sizerF)
#         self.Layout()
#
#     def __del__(self):
#         pass


class AdminUI(wx.Frame):
    
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, id=wx.ID_ANY, title=title, pos=wx.DefaultPosition,
                          size=wx.Size(800, 600), style=wx.DEFAULT_FRAME_STYLE | wx.TAB_TRAVERSAL)
        
        self.SetSizeHints(wx.DefaultSize, wx.DefaultSize)
        self.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_HIGHLIGHT))
        
        self._notebook = wx.Notebook(self, wx.ID_ANY, wx.DefaultPosition, wx.Size(800, 600), 0)
        self._notebook.SetFont(
            wx.Font(16, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString))
        self._notebook.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_3DLIGHT))
        
        tab_users = TabUsers(self._notebook)
        tab_books = TabBooks(self._notebook)
        # tab_statistics = TabStatistics(self._notebook)
        self._notebook.AddPage(tab_users, u"　用户　")
        self._notebook.AddPage(tab_books, u"　图书　")
        # self._notebook.AddPage(tab_statistics, u"　其他　")
        
        b_sizer = wx.BoxSizer(wx.VERTICAL)
        b_sizer.Add(self._notebook, 1, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
        self.SetSizerAndFit(b_sizer)
        
        self.Layout()
        
        self.Centre()
        self.Show()
    
    def __del__(self):
        pass


if __name__ == '__main__':
    app = wx.App()
    frame = AdminUI(None, u"图书管理员")
    app.MainLoop()
